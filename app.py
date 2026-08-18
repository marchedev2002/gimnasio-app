import os
import time
import calendar
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, send_file 
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash
import mysql.connector
from mysql.connector import pooling
from datetime import date, timedelta, datetime
from config import DB_CONFIG, SECRET_KEY, NOMBRE_GIMNASIO
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

app = Flask(__name__)
intentos_fallidos = {}
app.secret_key = SECRET_KEY
app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
@app.context_processor
def inject_nombre_gimnasio():
    """Inyecta el nombre del gimnasio en todas las plantillas."""
    nombre = session.get('nombre_gimnasio_sesion', NOMBRE_GIMNASIO)
    return dict(nombre_gimnasio=nombre)

# Carpeta donde se guardan las fotos de los socios
CARPETA_FOTOS = os.path.join(app.root_path, 'static', 'fotos')
EXTENSIONES_PERMITIDAS = {'jpg', 'jpeg', 'png'}


def extension_valida(nombre_archivo):
    """Chequea que el archivo subido sea una imagen permitida."""
    return '.' in nombre_archivo and \
           nombre_archivo.rsplit('.', 1)[1].lower() in EXTENSIONES_PERMITIDAS


def guardar_foto(dni, archivo):
    """
    Guarda la foto subida usando el DNI como nombre de archivo,
    para que siempre quede asociada al socio correcto.
    Devuelve el nombre de archivo guardado, o None si no se subió nada.
    """
    if archivo and archivo.filename != '' and extension_valida(archivo.filename):
        extension = archivo.filename.rsplit('.', 1)[1].lower()
        nombre_archivo = secure_filename(f"{dni}.{extension}")
        ruta_completa = os.path.join(CARPETA_FOTOS, nombre_archivo)
        archivo.save(ruta_completa)
        return nombre_archivo
    return None

def calcular_siguiente_vencimiento(fecha_base, dia_vencimiento):
    """
    Calcula la próxima fecha de vencimiento: un mes después de fecha_base,
    siempre anclada al mismo día del mes (dia_vencimiento), sin importar
    qué día se haya efectuado el pago real.
    """
    mes = fecha_base.month + 1
    anio = fecha_base.year
    if mes > 12:
        mes = 1
        anio += 1
    ultimo_dia_del_mes = calendar.monthrange(anio, mes)[1]
    dia = min(dia_vencimiento, ultimo_dia_del_mes)  # por si el día no existe en ese mes (ej: 31 en febrero)
    return date(anio, mes, dia)



pool_conexiones = None

def get_connection():
    global pool_conexiones

    for intento in range(3):
        try:
            if pool_conexiones is None:
                pool_conexiones = pooling.MySQLConnectionPool(
                    pool_name="gimnasio_pool",
                    pool_size=5,
                    **DB_CONFIG
                )
            return pool_conexiones.get_connection()
        except mysql.connector.Error:
            if intento < 2:
                time.sleep(3)  # le da tiempo a Aiven para "despertarse"
                pool_conexiones = None  # fuerza a recrear el pool en el próximo intento
            else:
                raise  # si después de 3 intentos sigue fallando, ahí sí mostramos el error

def login_requerido(vista):
    """Decorador que bloquea el acceso a una ruta si no hay sesion iniciada"""
    @wraps(vista)
    def envoltura(*args, **kwargs):
        if 'usuario' not in session:
            return redirect(url_for('login'))
        return vista(*args, **kwargs)
    return envoltura

@app.route('/')
@login_requerido
def index():
    """Página inicial: dashboard con buscador de DNI y estadísticas generales."""
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    hoy = date.today()
    cursor.execute("""
        SELECT COUNT(*) AS total,
               SUM(CASE WHEN fecha_proximo_vencimiento IS NOT NULL
                        AND fecha_proximo_vencimiento >= %s THEN 1 ELSE 0 END) AS al_dia
        FROM USUARIO
    """, (hoy,))
    stats = cursor.fetchone()

    cursor.close()
    conn.close()

    total_socios = stats['total'] or 0
    total_al_dia = stats['al_dia'] or 0
    total_vencidos = total_socios - total_al_dia

    return render_template(
        'index.html',
        total_socios=total_socios,
        total_al_dia=total_al_dia,
        total_vencidos=total_vencidos
    )

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Pagina de inicio de sesion para el personal"""
    if request.method == 'GET':
        return render_template('login.html')
    usuario_ingresado = request.form.get('usuario', '').strip()
    password_ingresado = request.form.get('password', '')

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    ip_origen = request.remote_addr
    intentos = intentos_fallidos.get(ip_origen, 0)

    if intentos >= 5:
        return render_template('login.html', error="Demasiados intentos fallidos. Esperá unos minutos e intentá de nuevo.")
    cursor.execute("""
        SELECT PERSONAL.*, GIMNASIO.nombre AS nombre_gimnasio
        FROM PERSONAL
        JOIN GIMNASIO ON PERSONAL.id_gimnasio = GIMNASIO.id_gimnasio
    WHERE PERSONAL.usuario = %s
    """, (usuario_ingresado,))
    personal = cursor.fetchone()
    cursor.close()
    conn.close()

    if personal and check_password_hash(personal['password_hash'], password_ingresado):
        session['usuario'] = personal['usuario']
        session['nombre'] = personal['nombre']
        session['id_gimnasio'] = personal['id_gimnasio']
        session['nombre_gimnasio_sesion'] = personal['nombre_gimnasio']
        return redirect(url_for('index'))
    else:
        intentos_fallidos[ip_origen] = intentos_fallidos.get(ip_origen, 0) + 1

    return render_template('login.html', error="Usuario o contraseña incorrectos.")

@app.route('/logout')
@login_requerido
def logout():
    """Cierra la sesion del personal"""
    session.clear()
    return redirect(url_for('login'))

@app.route('/buscar', methods=['POST'])
@login_requerido
def buscar():
    """Recibe el DNI por teclado (formulario) y redirige a la ficha del socio."""
    dni = request.form.get('dni', '').strip()

    if not dni:
        return render_template('index.html', error="Por favor ingresá un DNI.")

    return redirect(url_for('ver_socio', dni=dni))

@app.route('/nuevo', methods=['GET', 'POST'])
@login_requerido
def nuevo_socio():
    """Formulario para dar de alta un nuevo socio."""
    if request.method == 'GET':
        return render_template('form_socio.html', modo='nuevo', usuario=None)

    dni = request.form.get('dni', '').strip()
    nombre = request.form.get('nombre', '').strip()
    apellido = request.form.get('apellido', '').strip()
    telefono = request.form.get('telefono', '').strip()
    email = request.form.get('email', '').strip()

    if not dni or not nombre or not apellido:
        return render_template(
            'form_socio.html',
            modo='nuevo',
            usuario=None,
            error="DNI, nombre y apellido son obligatorios."
        )

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM USUARIO WHERE dni = %s AND id_gimnasio = %s", (dni, session['id_gimnasio']))
    if cursor.fetchone():
        cursor.close()
        conn.close()
        return render_template(
            'form_socio.html',
            modo='nuevo',
            usuario=None,
            error="Ya existe un socio con ese DNI."
        )

    nombre_foto = guardar_foto(dni, request.files.get('foto'))

    dia_vencimiento_nuevo = int(request.form.get('dia_vencimiento', date.today().day))

    cursor.execute("""
        INSERT INTO USUARIO (dni, nombre, apellido, telefono, email, foto, dia_vencimiento, fecha_proximo_vencimiento, id_gimnasio)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (dni, nombre, apellido, telefono or None, email or None, nombre_foto, dia_vencimiento_nuevo, None, session['id_gimnasio']))

    conn.commit()
    cursor.close()
    conn.close()

    return redirect(url_for('ver_socio', dni=dni, mensaje="Socio dado de alta correctamente."))

@app.route('/editar/<dni>', methods=['GET', 'POST'])
@login_requerido
def editar_socio(dni):
    """Formulario para editar los datos de un socio existente."""
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'GET':
        cursor.execute("SELECT * FROM USUARIO WHERE dni = %s AND id_gimnasio = %s", (dni, session['id_gimnasio']))
        usuario = cursor.fetchone()
        cursor.close()
        conn.close()

        if not usuario:
            return render_template('index.html', error=f"No se encontró ningún socio con DNI {dni}.")

        return render_template('form_socio.html', modo='editar', usuario=usuario)

    # POST: procesar la edición
    nombre = request.form.get('nombre', '').strip()
    apellido = request.form.get('apellido', '').strip()
    telefono = request.form.get('telefono', '').strip()
    email = request.form.get('email', '').strip()

    if not nombre or not apellido:
        cursor.execute("SELECT * FROM USUARIO WHERE dni = %s AND id_gimnasio = %s", (dni, session['id_gimnasio']))
        usuario = cursor.fetchone()
        cursor.close()
        conn.close()
        return render_template(
            'form_socio.html', modo='editar', usuario=usuario,
            error="Nombre y apellido son obligatorios."
        )

    # Si subieron una foto nueva, la guardamos y actualizamos el campo
    nombre_foto_nuevo = guardar_foto(dni, request.files.get('foto'))

    if nombre_foto_nuevo:
        cursor.execute("""
            UPDATE USUARIO
            SET nombre = %s, apellido = %s, telefono = %s, email = %s, foto = %s
            WHERE dni = %s AND id_gimnasio = %s
        """, (nombre, apellido, telefono or None, email or None, nombre_foto_nuevo, dni, session['id_gimnasio']))
    else:
        cursor.execute("""
            UPDATE USUARIO
            SET nombre = %s, apellido = %s, telefono = %s, email = %s
            WHERE dni = %s
        """, (nombre, apellido, telefono or None, email or None, dni))

    conn.commit()
    cursor.close()
    conn.close()

    return redirect(url_for('ver_socio', dni=dni, mensaje='edicion'))


@app.route('/socio/<dni>')
@login_requerido
def ver_socio(dni):
    """Muestra la ficha de un socio, usando su día de vencimiento fijo."""
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM USUARIO WHERE dni = %s AND id_gimnasio = %s", (dni, session['id_gimnasio']))
    usuario = cursor.fetchone()

    if not usuario:
        cursor.close()
        conn.close()
        return render_template('index.html', error=f"No se encontró ningún socio con DNI {dni}.")

    hoy = date.today()
    membresia_al_dia = (
        usuario['fecha_proximo_vencimiento'] is not None
        and usuario['fecha_proximo_vencimiento'] >= hoy
    )

    cursor.execute("""
        SELECT PAGO.id_pago, PAGO.fecha_pago, MES.nombre_mes, PAGO.anio,
               PRECIO.tipo_membresia, PRECIO.monto
        FROM PAGO
        JOIN MES ON PAGO.id_mes = MES.id_mes
        JOIN PRECIO ON PAGO.id_precio = PRECIO.id_precio
        WHERE PAGO.dni = %s
        ORDER BY PAGO.fecha_pago DESC
    """, (dni,))
    historial_pagos = cursor.fetchall()

    cursor.close()
    conn.close()

    mensaje = request.args.get('mensaje')

    return render_template(
        'resultado.html',
        usuario=usuario,
        membresia_al_dia=membresia_al_dia,
        historial_pagos=historial_pagos,
        mensaje=mensaje
    )

@app.route('/pago/nuevo', methods=['GET', 'POST'])
@login_requerido
def nuevo_pago():
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'GET':
        cursor.execute("SELECT dni, nombre, apellido FROM USUARIO WHERE id_gimnasio = %s ORDER BY apellido, nombre", (session['id_gimnasio'],))
        socios = cursor.fetchall()
        cursor.execute("SELECT id_mes, nombre_mes FROM MES ORDER BY id_mes")
        meses = cursor.fetchall()
        cursor.execute("SELECT id_precio, tipo_membresia, monto FROM PRECIO WHERE activo = 1 AND id_gimnasio = %s ORDER BY tipo_membresia", (session['id_gimnasio'],))
        precios = cursor.fetchall()
        cursor.close()
        conn.close()

        hoy = date.today()
        dni_preseleccionado = request.args.get('dni', '')

        return render_template(
            'form_pago.html', socios=socios, meses=meses, precios=precios,
            dni_preseleccionado=dni_preseleccionado, mes_actual=hoy.month,
            anio_actual=hoy.year, fecha_hoy=hoy.isoformat()
        )

    dni = request.form.get('dni', '').strip()
    id_mes = request.form.get('id_mes', '').strip()
    anio = request.form.get('anio', '').strip()
    id_precio = request.form.get('id_precio', '').strip()
    fecha_pago = request.form.get('fecha_pago', '').strip()
    metodo_pago = request.form.get('metodo_pago', '').strip()

    if not dni or not id_mes or not anio or not id_precio or not fecha_pago or not metodo_pago:
        cursor.execute("SELECT dni, nombre, apellido FROM USUARIO WHERE id_gimnasio = %s ORDER BY apellido, nombre", (session['id_gimnasio'],))
        socios = cursor.fetchall()
        cursor.execute("SELECT id_mes, nombre_mes FROM MES ORDER BY id_mes")
        meses = cursor.fetchall()
        cursor.execute("SELECT id_precio, tipo_membresia, monto FROM PRECIO WHERE activo = 1 AND id_gimnasio = %s ORDER BY tipo_membresia", (session['id_gimnasio'],))
        precios = cursor.fetchall()
        cursor.close()
        conn.close()

        hoy = date.today()
        return render_template(
            'form_pago.html', socios=socios, meses=meses, precios=precios,
            dni_preseleccionado=dni, mes_actual=hoy.month, anio_actual=hoy.year,
            fecha_hoy=hoy.isoformat(), error="Completá todos los campos, incluido el método de pago."
        )

    cursor.execute("""
        INSERT INTO PAGO (dni, id_mes, anio, id_precio, fecha_pago, metodo_pago, id_gimnasio)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (dni, id_mes, anio, id_precio, fecha_pago, metodo_pago, session['id_gimnasio']))

    cursor.execute("SELECT dia_vencimiento FROM USUARIO WHERE dni = %s AND id_gimnasio = %s", (dni, session['id_gimnasio']))
    socio = cursor.fetchone()
    dia_vto = socio['dia_vencimiento']

    id_mes_int = int(id_mes)
    anio_int = int(anio)

    ultimo_dia_mes_pagado = calendar.monthrange(anio_int, id_mes_int)[1]
    dia_ajustado = min(dia_vto, ultimo_dia_mes_pagado)
    fecha_periodo_pagado = date(anio_int, id_mes_int, dia_ajustado)

    nueva_fecha_vencimiento = calcular_siguiente_vencimiento(fecha_periodo_pagado, dia_vto)
    cursor.execute(
        "UPDATE USUARIO SET fecha_proximo_vencimiento = %s WHERE dni = %s AND id_gimnasio = %s",
        (nueva_fecha_vencimiento, dni, session['id_gimnasio'])
    )
    conn.commit()
    cursor.close()
    conn.close()

    return redirect(url_for('ver_socio', dni=dni, mensaje='pago'))

@app.route('/socios')
@login_requerido
def listado_socios():
    """Muestra el listado completo de todos los socios con su estado y filtros"""
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    hoy = date.today()

    cursor.execute("""
        SELECT dni, nombre, apellido, fecha_proximo_vencimiento,
            CASE WHEN fecha_proximo_vencimiento IS NOT NULL AND fecha_proximo_vencimiento >= %s
                    THEN 1 ELSE 0 END AS al_dia
        FROM USUARIO
        WHERE id_gimnasio = %s
        ORDER BY apellido, nombre
    """, (hoy, session['id_gimnasio']))

    todos = cursor.fetchall()
    cursor.close()
    conn.close()

    total_al_dia = sum(1 for s in todos if s['al_dia'] == 1)
    total_vencidos = len(todos) - total_al_dia

    busqueda = request.args.get('buscar', '').strip().lower()
    filtro = request.args.get('filtro', 'todos')

    socios = todos

    if busqueda:
        socios = [
            s for s in socios
            if busqueda in s['nombre'].lower() or busqueda in s['apellido'].lower() or busqueda in s['dni'].lower()
        ]

    if filtro == 'al_dia':
        socios = [s for s in socios if s['al_dia'] == 1]
    elif filtro == 'vencidos':
        socios = [s for s in socios if s['al_dia'] == 0]

    return render_template(
        'listado.html',
        socios=socios,
        total_al_dia=total_al_dia,
        total_vencidos=total_vencidos,
        busqueda=busqueda,
        filtro=filtro
    )


@app.route('/eliminar/<dni>', methods=['GET','POST'])
@login_requerido
def eliminar_socio(dni):
    """Elmina un socio y sus pagos de la base"""
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM USUARIO WHERE dni = %s AND id_gimnasio = %s", (dni, session['id_gimnasio']))
    usuario = cursor.fetchone()

    if not usuario:
        cursor.close()
        conn.close()
        return render_template('index.html', error=f"No se encontró ningún socio con DNI {dni}.")

    if request.method == 'GET':
        cursor.close()
        conn.close()
        return render_template('confirmar_eliminar.html', usuario=usuario)

    cursor.execute("DELETE FROM ASISTENCIA WHERE dni = %s AND id_gimnasio = %s", (dni, session['id_gimnasio']))
    cursor.execute("DELETE FROM PAGO WHERE dni = %s AND id_gimnasio = %s", (dni, session['id_gimnasio']))
    cursor.execute("DELETE FROM USUARIO WHERE dni = %s AND id_gimnasio = %s", (dni, session['id_gimnasio']))
    conn.commit()
    cursor.close()
    conn.close()

    return redirect(url_for('listado_socios', mensaje='baja'))

@app.route('/alertas')
@login_requerido
def alertas():
    """Muestra socios vencidos y por vencer, según el día fijo de cada uno."""
    dias_limite = request.args.get('dias', 5)
    try:
        dias_limite = int(dias_limite)
    except ValueError:
        dias_limite = 5

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT dni, nombre, apellido, fecha_proximo_vencimiento
        FROM USUARIO
        WHERE id_gimnasio = %s
        ORDER BY apellido, nombre
    """, (session['id_gimnasio'],))
    todos = cursor.fetchall()

    cursor.close()
    conn.close()

    hoy = date.today()

    vencidos = [
        s for s in todos
        if s['fecha_proximo_vencimiento'] is None or s['fecha_proximo_vencimiento'] < hoy
    ]

    por_vencer = [
        s for s in todos
        if s['fecha_proximo_vencimiento'] is not None
        and hoy <= s['fecha_proximo_vencimiento'] <= hoy + timedelta(days=dias_limite)
    ]

    return render_template(
        'alertas.html',
        vencidos=vencidos,
        por_vencer=por_vencer,
        dias_limite=dias_limite
    )

@app.route('/reportes')
@login_requerido
def reportes():
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    hoy = date.today()
    inicio_semana = hoy - timedelta(days=hoy.weekday())  # lunes de esta semana
    inicio_mes = date(hoy.year, hoy.month, 1)

    def resumen_desde(fecha_desde):
        cursor.execute("""
            SELECT PAGO.metodo_pago, SUM(PRECIO.monto) AS total, COUNT(*) AS cantidad
            FROM PAGO
            JOIN PRECIO ON PAGO.id_precio = PRECIO.id_precio
            WHERE PAGO.fecha_pago >= %s AND PAGO.id_gimnasio = %s
            GROUP BY PAGO.metodo_pago
        """, (fecha_desde, session['id_gimnasio']))
        filas = cursor.fetchall()

        total = sum(float(f['total']) for f in filas)
        cantidad = sum(f['cantidad'] for f in filas)
        efectivo = next((float(f['total']) for f in filas if f['metodo_pago'] == 'Efectivo'), 0)
        debito = next((float(f['total']) for f in filas if f['metodo_pago'] == 'Debito'), 0)

        return {'total': total, 'cantidad': cantidad, 'efectivo': efectivo, 'debito': debito}

    resumen_dia = resumen_desde(hoy)
    resumen_semana = resumen_desde(inicio_semana)
    resumen_mes = resumen_desde(inicio_mes)

    # --- Distribución de ingresos por profesor (para el gráfico de torta) ---
    meses_grafico = request.args.get('meses', 1)
    try:
        meses_grafico = int(meses_grafico)
    except ValueError:
        meses_grafico = 1
    meses_grafico = max(1, min(12, meses_grafico))  # entre 1 y 12

    # Calcular desde qué fecha mirar hacia atrás (ej: si meses_grafico=3, arranca
    # el día 1 del mes de hace 2 meses, incluyendo el mes actual)
    mes_desde = hoy.month - (meses_grafico - 1)
    anio_desde = hoy.year
    while mes_desde <= 0:
        mes_desde += 12
        anio_desde -= 1
    fecha_desde_grafico = date(anio_desde, mes_desde, 1)

    cursor.execute("""
        SELECT COALESCE(CONCAT(PROFESOR.nombre, ' ', PROFESOR.apellido), 'Musculación libre / Sin clase') AS categoria,
            COUNT(*) AS cantidad
        FROM ASISTENCIA
        LEFT JOIN CLASE ON ASISTENCIA.id_clase = CLASE.id_clase
        LEFT JOIN PROFESOR ON CLASE.id_profesor = PROFESOR.id_profesor
        WHERE ASISTENCIA.fecha_hora >= %s AND ASISTENCIA.id_gimnasio = %s
        GROUP BY categoria
        ORDER BY cantidad DESC
    """, (fecha_desde_grafico, session['id_gimnasio']))
    distribucion_profesores = cursor.fetchall()

    total_ingresos_periodo = sum(fila['cantidad'] for fila in distribucion_profesores)
    for fila in distribucion_profesores:
        fila['porcentaje'] = round((fila['cantidad'] / total_ingresos_periodo * 100), 1) if total_ingresos_periodo else 0

    #Tendencia de ingresos: ultimos 12 meses
    fecha_hace_12_meses = hoy.replace(day=1)
    for _ in range(11):
        fecha_hace_12_meses = (fecha_hace_12_meses - timedelta(days=1)).replace(day=1)

    cursor.execute("""
        SELECT DATE_FORMAT(fecha_pago, '%Y-%m') AS periodo, SUM(PRECIO.monto) AS total
        FROM PAGO
        JOIN PRECIO ON PAGO.id_precio = PRECIO.id_precio
        WHERE fecha_pago >= %s AND PAGO.id_gimnasio = %s
        GROUP BY periodo
        ORDER BY periodo
    """, (fecha_hace_12_meses, session['id_gimnasio']))
    filas_tendencia = cursor.fetchall()

    #Rellenamos los meses sin pagos con $0
    mapa_totales = {fila['periodo']: float(fila['total']) for fila in filas_tendencia}
    tendencia_labels = []
    tendencia_valores = []
    cursor_mes = fecha_hace_12_meses
    for _ in range(12):
        clave = cursor_mes.strftime('%Y-%m')
        tendencia_labels.append(cursor_mes.strftime('%b %Y'))
        tendencia_valores.append(mapa_totales.get(clave, 0))
        cursor_mes = date(cursor_mes.year + (1 if cursor_mes.month == 12 else 0),
                          1 if cursor_mes.month == 12 else cursor_mes.month + 1, 1)

    #Porcentaje de Retencion: % de socios que renuevan mes a mes, los ultimos 6 periodos
    periodos = []
    anio_iter, mes_iter = hoy.year, hoy.month
    for _ in range(7):
        periodos.insert(0, (anio_iter, mes_iter))
        mes_iter -= 1
        if mes_iter == 0:
            mes_iter == 12
            anio_iter -=1

    anio_min, mes_min = periodos[0]
    cursor.execute("""
        SELECT DISTINCT dni, id_mes, anio FROM PAGO
        WHERE ((anio > %s) OR (anio = %s AND id_mes >= %s)) AND id_gimnasio = %s
    """, (anio_min, anio_min, mes_min, session['id_gimnasio']))
    pagos_periodo = cursor.fetchall()

    socios_por_periodo = {}
    for fila in pagos_periodo:
        clave = (fila['anio'], fila['id_mes'])
        socios_por_periodo.setdefault(clave, set()).add(fila['dni'])

    nombres_meses = ['', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    retencion_labels = []
    retencion_valores = []
    for i in range(1, 7):
        periodo_anterior = periodos[i - 1]
        periodo_actual = periodos[i]
        base = socios_por_periodo.get(periodo_anterior, set())
        actuales = socios_por_periodo.get(periodo_actual, set())
        renovaron = base & actuales

        tasa = round((len(renovaron) / len(base) * 100), 1) if base else None
        retencion_labels.append(f"{nombres_meses[periodo_actual[1]]} {periodo_actual[0]}")
        retencion_valores.append(tasa)


    cursor.execute("SELECT id_mes, nombre_mes FROM MES ORDER BY id_mes")
    meses = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        'reportes.html',
        resumen_dia=resumen_dia, resumen_semana=resumen_semana, resumen_mes=resumen_mes,
        meses=meses, mes_actual=hoy.month, anio_actual=hoy.year, distribucion_profesores = distribucion_profesores, meses_grafico=meses_grafico,
        total_ingresos_periodo = total_ingresos_periodo,
        tendencia_labels=tendencia_labels,
        tendencia_valores=tendencia_valores,
        retencion_labels=retencion_labels,
        retencion_valores=retencion_valores
    )

@app.route('/admin')
@login_requerido
def admin_general():
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM PRECIO WHERE activo = TRUE AND id_gimnasio = %s ORDER BY tipo_membresia", (session['id_gimnasio'],))
    precios = cursor.fetchall()

    cursor.execute("SELECT * FROM PROFESOR WHERE id_gimnasio = %s ORDER BY apellido, nombre", (session['id_gimnasio'],))
    profesores = cursor.fetchall()

    cursor.execute("""
        SELECT CLASE.*, PROFESOR.nombre AS profesor_nombre, PROFESOR.apellido AS profesor_apellido
        FROM CLASE
        JOIN PROFESOR ON CLASE.id_profesor = PROFESOR.id_profesor
        WHERE CLASE.id_gimnasio = %s
        ORDER BY CLASE.hora_inicio
    """, (session['id_gimnasio'],))
    clases = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        'admin.html',
        precios=precios,
        profesores=profesores,
        clases=clases)

@app.route('/admin/precio/nuevo', methods=['GET', 'POST'])
@login_requerido
def nuevo_precio():
    if request.method == 'GET':
        return render_template('form_precio.html')

    tipo_membresia = request.form.get('tipo_membresia', '').strip()
    monto = request.form.get('monto', '').strip()
    dias_max_mes = request.form.get('dias_max_mes', '').strip()

    if not tipo_membresia or not monto:
        return render_template('form_precio.html', error="Completá el nombre y el monto.")

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO PRECIO (tipo_membresia, monto, dias_max_mes, activo, id_gimnasio)
        VALUES (%s, %s, %s, TRUE, %s)
    """, (tipo_membresia, monto, dias_max_mes or None, session['id_gimnasio']))
    conn.commit()
    cursor.close()
    conn.close()

    return redirect(url_for('admin_general'))


@app.route('/admin/precio/editar/<int:id_precio>', methods=['POST'])
@login_requerido
def editar_precio(id_precio):
    nuevo_monto = request.form.get('monto', '').strip()

    if not nuevo_monto:
        return redirect(url_for('admin_general'))

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT tipo_membresia, dias_max_mes FROM PRECIO WHERE id_precio = %s AND id_gimnasio = %s", (id_precio, session['id_gimnasio']))
    precio_actual = cursor.fetchone()

    if precio_actual:
        cursor.execute("UPDATE PRECIO SET activo = FALSE WHERE id_precio = %s AND id_gimnasio = %s", (id_precio, session['id_gimnasio']))
        cursor.execute("""
            INSERT INTO PRECIO (tipo_membresia, monto, dias_max_mes, activo, id_gimnasio)
            VALUES (%s, %s, %s, TRUE, %s)
        """, (precio_actual['tipo_membresia'], nuevo_monto, precio_actual['dias_max_mes'], session['id_gimnasio']))
        conn.commit()

    cursor.close()
    conn.close()

    return redirect(url_for('admin_general'))


@app.route('/admin/profesor/nuevo', methods=['GET', 'POST'])
@login_requerido
def nuevo_profesor():
    if request.method == 'GET':
        return render_template('form_profesor.html')

    nombre = request.form.get('nombre', '').strip()
    apellido = request.form.get('apellido', '').strip()

    if not nombre or not apellido:
        return render_template('form_profesor.html', error="Completá nombre y apellido.")

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO PROFESOR (nombre, apellido, id_gimnasio) VALUES (%s, %s, %s)", (nombre, apellido, session['id_gimnasio']))
    conn.commit()
    cursor.close()
    conn.close()

    return redirect(url_for('admin_general'))


@app.route('/admin/clase/nueva', methods=['GET', 'POST'])
@login_requerido
def nueva_clase():
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'GET':
        cursor.execute("SELECT id_profesor, nombre, apellido FROM PROFESOR WHERE id_gimnasio = %s ORDER BY apellido, nombre", (session['id_gimnasio'],))
        profesores = cursor.fetchall()
        cursor.close()
        conn.close()
        return render_template('form_clase.html', profesores=profesores)

    nombre = request.form.get('nombre', '').strip()
    hora_inicio = request.form.get('hora_inicio', '').strip()
    hora_fin = request.form.get('hora_fin', '').strip()
    id_profesor = request.form.get('id_profesor', '').strip()

    if not nombre or not hora_inicio or not hora_fin or not id_profesor:
        cursor.execute("SELECT id_profesor, nombre, apellido FROM PROFESOR WHERE id_gimnasio = %s ORDER BY apellido, nombre", (session['id_gimnasio'],))
        profesores = cursor.fetchall()
        cursor.close()
        conn.close()
        return render_template('form_clase.html', profesores=profesores, error="Completá todos los campos.")

    cursor.execute("""
        INSERT INTO CLASE (nombre, hora_inicio, hora_fin, id_profesor, id_gimnasio)
        VALUES (%s, %s, %s, %s, %s)
    """, (nombre, hora_inicio, hora_fin, id_profesor, session['id_gimnasio']))
    conn.commit()
    cursor.close()
    conn.close()

    return redirect(url_for('admin_general'))

@app.route('/reportes/exportar')
@login_requerido
def exportar_reporte():
    """Genera un Excel con todos los pagos del mes/año seleccionado."""
    mes = request.args.get('mes', date.today().month, type=int)
    anio = request.args.get('anio', date.today().year, type=int)

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT usuario.dni, usuario.nombre, usuario.apellido,
            mes.nombre_mes, pago.anio, precio.tipo_membresia, precio.monto, pago.fecha_pago
        FROM PAGO
        JOIN USUARIO ON PAGO.dni = USUARIO.dni AND PAGO.id_gimnasio = USUARIO.id_gimnasio
        JOIN MES ON PAGO.id_mes = MES.id_mes
        JOIN PRECIO ON PAGO.id_precio = PRECIO.id_precio
        WHERE PAGO.id_mes = %s AND PAGO.anio = %s AND PAGO.id_gimnasio = %s
        ORDER BY USUARIO.apellido, USUARIO.nombre
    """, (mes, anio, session['id_gimnasio']))
    pagos = cursor.fetchall()
    cursor.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.tittle = "Pagos"

    encabezados = ['DNI', 'Nombre', 'Apellido', 'Mes', 'Año', 'Plan', 'Monto', 'Fecha de pago']
    ws.append(encabezados)

    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="E94560", end_color="E94560", fill_type="solid")

    total = 0
    for p in pagos:
        ws.append([
            p['dni'], p['nombre'], p['apellido'], p['nombre_mes'],
            p['anio'], p['tipo_membresia'], float(p['monto']), p['fecha_pago'].strftime('d/%m/%Y')
        ])
        total += float(p['monto'])

    ws.append([])
    ws.append(['', '', '', '', '', 'TOTAL RECAUDADO', total,'' ])
    ws.cell(row=ws.max_row, column=6).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=7).font = Font(bold=True)

    # Ajustar ancho de columnas
    for columna in ws.columns:
        ancho_max = max(len(str(celda.value)) if celda.value else 0 for celda in columna)
        ws.column_dimensions[columna[0].column_letter].width = ancho_max + 3

    archivo_en_memoria = BytesIO()
    wb.save(archivo_en_memoria)
    archivo_en_memoria.seek(0)

    nombre_mes_texto = next((m for m in ['', 'Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio',
                                          'Agosto','Septiembre','Octubre','Noviembre','Diciembre']
                              if False), None)  # placeholder, se resuelve abajo

    nombre_archivo = f"pagos_{mes:02d}-{anio}.xlsx"

    return send_file(
        archivo_en_memoria,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


    if not mes or not anio:
        return redirect(url_for('reportes', error="Por favor seleccioná un mes y un año."))

@app.route('/checkin', methods=['GET'])
@login_requerido
def checkin():
    """Pantalla rápida de ingreso: solo pide el DNI."""
    return render_template('checkin.html')


@app.route('/checkin/procesar', methods=['POST'])
@login_requerido
def procesar_checkin():
    """Registra el ingreso del socio y muestra su estado + visitas del mes."""
    dni = request.form.get('dni', '').strip()

    if not dni:
        return render_template('checkin.html', error="Por favor ingresá un DNI.")

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM USUARIO WHERE dni = %s AND id_gimnasio = %s", (dni, session['id_gimnasio']))
    usuario = cursor.fetchone()

    if not usuario:
        cursor.close()
        conn.close()
        return render_template('checkin.html', error=f"No se encontró ningún socio con DNI {dni}.")

    # Registrar el ingreso (fecha y hora actuales)
    ahora = datetime.now()
    hora_actual = ahora.time()
    cursor.execute("""
        SELECT id_clase FROM CLASE
        WHERE hora_inicio <= %s AND hora_fin >= %s AND id_gimnasio = %s
        LIMIT 1
    """, (hora_actual, hora_actual, session['id_gimnasio']))
    clase_en_curso = cursor.fetchone()
    id_clase_actual = clase_en_curso['id_clase'] if clase_en_curso else None

    cursor.execute(
    "INSERT INTO ASISTENCIA (dni, fecha_hora, id_clase, id_gimnasio) VALUES (%s, %s, %s, %s)",
    (dni, ahora, id_clase_actual, session['id_gimnasio'])
    )
    conn.commit()

    # Estado de la membresía (misma lógica que ya usábamos en ver_socio)
    hoy = date.today()
    membresia_al_dia = (
        usuario['fecha_proximo_vencimiento'] is not None
        and usuario['fecha_proximo_vencimiento'] >= hoy
    )

    # Plan actual del socio: el de su último pago registrado
    cursor.execute("""
        SELECT PRECIO.tipo_membresia, PRECIO.dias_max_mes
        FROM PAGO
        JOIN PRECIO ON PAGO.id_precio = PRECIO.id_precio
        WHERE PAGO.dni = %s AND PAGO.id_gimnasio = %s
        ORDER BY PAGO.fecha_pago DESC
        LIMIT 1
    """, (dni, session['id_gimnasio']))
    plan_actual = cursor.fetchone()

    # Cuántas veces ya asistió este mes calendario
    cursor.execute("""
        SELECT COUNT(*) AS cantidad
        FROM ASISTENCIA
        WHERE dni = %s AND id_gimnasio = %s
        AND EXTRACT(YEAR FROM fecha_hora) = EXTRACT(YEAR FROM CURRENT_DATE)
        AND EXTRACT(MONTH FROM fecha_hora) = EXTRACT(MONTH FROM CURRENT_DATE)
    """, (dni, session['id_gimnasio']))
    visitas_mes = cursor.fetchone()['cantidad']

    cursor.close()
    conn.close()

    excede_limite = (
        plan_actual is not None
        and plan_actual['dias_max_mes'] is not None
        and visitas_mes > plan_actual['dias_max_mes']
    )

    return render_template(
        'resultado_checkin.html',
        usuario=usuario,
        membresia_al_dia=membresia_al_dia,
        hora_ingreso=ahora.strftime('%H:%M'),
        plan_actual=plan_actual,
        visitas_mes=visitas_mes,
        excede_limite=excede_limite
    )


if __name__ == '__main__':
    app.run(debug=False)
